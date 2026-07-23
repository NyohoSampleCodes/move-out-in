import openpyxl

workbook = openpyxl.load_workbook("data/raw/idou_2024.xlsx", data_only=True)
sheet = workbook["a002"]

value = sheet.cell(row=9, column=7).value
print(value)

value = sheet.cell(row=9, column=14).value
print(value, "人（北海道 → 青森県）")

value = sheet.cell(row=9, column=11).value
print(repr(value), type(value))

# 47都道府県 x 47都道府県、全マスを二重ループで読んでみる
# （行=転出元、9行目から。列=転入先、11列目から3列おき）
count = 0
dash_count = 0
for row_index in range(47):
    excel_row = 9 + row_index
    for col_index in range(47):
        excel_col = 11 + col_index * 3
        value = sheet.cell(row=excel_row, column=excel_col).value
        count += 1
        if value == "-":
            dash_count += 1

print("読んだセルの数:", count)
print("'-' だったセルの数:", dash_count, "（47都道府県ぶんの対角成分のはず）")
